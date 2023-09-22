from pathlib import Path
from PIL import Image
import random
import numpy as np
import torch

import openpyxl
import cv2
from scipy.ndimage.interpolation import rotate

from natsort import natsorted

import matplotlib.pyplot as plt #試し用


class Train_val(object):
    def __init__(self,train_img,train_mask,train_gt):
        self.img_path = train_img
        self.mask_path = train_mask
        self.gt = train_gt

    def __len__(self):
        return int(len(self.img_path))
    
    def __getitem__(self,idx):
        img_path = self.img_path[idx]
        img_path = str(img_path)
        img = cv2.imread(img_path)

        mask_path = self.mask_path[idx]
        mask_path = str(mask_path)
        mask = cv2.imread(mask_path)

        rand1 = random.randint(0,3)
        rand2 = random.randint(-2,1)
        if rand1 == 1:
            img = cv2.rotate(img, cv2.ROTATE_90_CLOCKWISE)
            mask = cv2.rotate(mask, cv2.ROTATE_90_CLOCKWISE)
        elif rand1 == 2:
            img = cv2.rotate(img, cv2.ROTATE_180)
            mask = cv2.rotate(mask, cv2.ROTATE_180)
        elif rand1 == 3:
            img = cv2.rotate(img, cv2.ROTATE_90_COUNTERCLOCKWISE)
            mask = cv2.rotate(mask, cv2.ROTATE_90_COUNTERCLOCKWISE)

        if rand2 >= -1:
            img = cv2.flip(img,rand2)
            mask = cv2.flip(mask,rand2)

        img = img / 255
        mask = mask / 255

        gt = self.gt[idx]
        
        img = torch.from_numpy(img.astype(np.float32))
        mask = torch.from_numpy(mask.astype(np.float32))
        return img.permute(2, 0, 1), img_path, mask.permute(2, 0, 1), gt, idx

class Pred_old(object):
    def __init__(self,img_path, mask_path, gt_path):
        self.img_paths = img_path
        self.mask_paths = mask_path
        self.gt = gt_path
    
    def __len__(self):
        return len(self.img_paths)

    def __getitem__(self,idx):
        img_path = self.img_paths[idx]
        img_path = str(img_path)
        img = cv2.imread(img_path)
        img = img / 255

        mask_path = self.mask_paths[idx]
        mask_path = str(mask_path)
        mask = cv2.imread(mask_path)
        mask = mask / 255

        gt = self.gt[idx]

        img = torch.from_numpy(img.astype(np.float32))
        mask = torch.from_numpy(mask.astype(np.float32))
        return img.permute(2, 0, 1), img_path, mask.permute(2, 0, 1), gt, idx


class Train_val_nonmask(object):
    def __init__(self,train_img,train_gt):
        self.img_path = train_img
        self.gt = train_gt

    def __len__(self):
        return len(self.img_path)
    
    def __getitem__(self,idx):
        img_path = self.img_path[idx]
        img_path = str(img_path)
        img = cv2.imread(img_path)


        rand1 = random.randint(0,3)
        rand2 = random.randint(-2,1)
        if rand1 == 1:
            img = cv2.rotate(img, cv2.ROTATE_90_CLOCKWISE)
        elif rand1 == 2:
            img = cv2.rotate(img, cv2.ROTATE_180)
        elif rand1 == 3:
            img = cv2.rotate(img, cv2.ROTATE_90_COUNTERCLOCKWISE)

        if rand2 >= -1:
            img = cv2.flip(img,rand2)

        img = img / 255

        gt = self.gt[idx]
        
        img = torch.from_numpy(img.astype(np.float32))
        return img.permute(2, 0, 1), img_path, gt, idx

class Pred_nonmask(object):
    def __init__(self,img_path):
        self.img_paths = img_path
    
    def __len__(self):
        return len(self.img_paths)

    def __getitem__(self,idx):
        img_path = self.img_paths[idx]
        img_path = str(img_path)
        img = cv2.imread(img_path)
        img = img / 255

        img = torch.from_numpy(img.astype(np.float32))
        return img.permute(2, 0, 1), img_path, idx


class Est_proportion_train(object):
    def __init__(self,base_path):
        self.img_paths0 = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_2_PDL122C3").glob("*.png"))
        self.mask_paths0 = sorted(Path(f"{base_path}/256_data/point_mask/96_2_PDL122C3").glob("*.png"))
        self.img_paths1 = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_3_PDL122C3").glob("*.png"))
        self.mask_paths1 = sorted(Path(f"{base_path}/256_data/point_mask/96_3_PDL122C3").glob("*.png"))
        self.img_paths2 = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_4_PDL122C3").glob("*.png"))
        self.mask_paths2 = sorted(Path(f"{base_path}/256_data/point_mask/96_4_PDL122C3").glob("*.png"))

        self.img_paths = self.img_paths0 + self.img_paths1 + self.img_paths2
        self.mask_paths = self.mask_paths0 + self.mask_paths1 + self.mask_paths2

        # over sampling
        # exel sheet 2 numpy array
        input_file_name = '/home/okuo-takumi/pdl1_code/pdl1_rate/eval/delete7/train/96_234_rate.xlsx'
        book = openpyxl.load_workbook(input_file_name)
        sheet = book['Sheet1']
        gt_labels = []
        for idx in range(sheet.max_row - 1):
            gt_val = sheet[f'H{idx+2}'].value
            if gt_val == 1:
                gt_val = 3
            if gt_val == 2:
                gt_val = 3
            gt_labels.append(gt_val)
        gt_labels = np.array(gt_labels)

        # max num of class
        img_path_np = np.array(self.img_paths)
        mask_path_np = np.array(self.mask_paths)

        label_num_list = []
        img_path_per_label = {}
        mask_path_per_label = {}
        for u_id in np.unique(gt_labels):
            label_num_list.append(sum(gt_labels == u_id))
            img_path_per_label[u_id] = img_path_np[gt_labels == u_id]
            mask_path_per_label[u_id] = mask_path_np[gt_labels == u_id]
        num_img = max(label_num_list)
        
        new_img_path = []
        new_mask_path = []
        new_gt = []

        for u_id, num_list in zip(np.unique(gt_labels), label_num_list):
            x = num_img // num_list
            for idx in range(x):
                new_img_path.extend(img_path_per_label[u_id])
            x = num_img % num_list
            new_img_path.extend(img_path_per_label[u_id][:x])

        for u_id, num_list in zip(np.unique(gt_labels), label_num_list):
            x = num_img // num_list
            for idx in range(x):
                new_mask_path.extend(mask_path_per_label[u_id])
            x = num_img % num_list
            new_mask_path.extend(mask_path_per_label[u_id][:x])

        for u_id, num_list in zip(np.unique(gt_labels), label_num_list):
            x = num_img // num_list
            for idx in range(x):
                new_gt.extend(gt_labels[gt_labels == u_id])
            x = num_img % num_list
            new_gt.extend(gt_labels[gt_labels == u_id][:x])

        self.new_img_path = new_img_path
        self.new_mask_path = new_mask_path
        self.new_gt = new_gt

        

    def __len__(self):
        return len(self.new_img_path)

    def __getitem__(self,idx):
        img_path = self.new_img_path[idx]
        img_path = str(img_path)
        img = cv2.imread(img_path)
        img = img / 255

        mask_path = self.new_mask_path[idx]
        mask_path = str(mask_path)
        mask = cv2.imread(mask_path)
        # mask_1lay = mask[:,:,0]
        mask = mask / 255
        # mask = cv2.resize(mask, dsize = None, fx = 0.25, fy = 0.25)
        # 試し用
        # plt.imshow(img)
        # if cpoint.shape != (0,):
        #     plt.scatter(cpoint[:,0],cpoint[:,1], c = "red")
        # if npoint.shape != (0,):    
        #     plt.scatter(npoint[:,0],npoint[:,1], c = "y")
        # plt.show()
        # plt.close()
        # ここまで

        gt = self.new_gt[idx]
        
        img = torch.from_numpy(img.astype(np.float32))
        mask = torch.from_numpy(mask.astype(np.float32))
        return img.permute(2, 0, 1), img_path, mask.permute(2, 0, 1), gt, idx
        # return img, img_path, mask

class Est_proportion_val(object):
    def __init__(self,base_path):
        self.img_paths0 = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_2_PDL122C3").glob("*.png"))
        self.mask_paths0 = sorted(Path(f"{base_path}/256_data/point_mask/96_2_PDL122C3").glob("*.png"))
        self.img_paths1 = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_3_PDL122C3").glob("*.png"))
        self.mask_paths1 = sorted(Path(f"{base_path}/256_data/point_mask/96_3_PDL122C3").glob("*.png"))
        self.img_paths2 = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_4_PDL122C3").glob("*.png"))
        self.mask_paths2 = sorted(Path(f"{base_path}/256_data/point_mask/96_4_PDL122C3").glob("*.png"))

        self.img_paths = self.img_paths0 + self.img_paths1 + self.img_paths2
        self.mask_paths = self.mask_paths0 + self.mask_paths1 + self.mask_paths2

        # over sampling
        # exel sheet 2 numpy array
        input_file_name = '/home/okuo-takumi/pdl1_code/pdl1_rate/eval/delete7/train/96_234_rate.xlsx'
        book = openpyxl.load_workbook(input_file_name)
        sheet = book['Sheet1']
        gt_labels = []
        for idx in range(sheet.max_row - 1):
            gt_val = sheet[f'H{idx+2}'].value
            if gt_val == 1:
                gt_val = 3
            if gt_val == 2:
                gt_val = 3
            gt_labels.append(gt_val)
        gt_labels = np.array(gt_labels)
        self.gt = gt_labels
        

    def __len__(self):
        return len(self.img_paths)

    def __getitem__(self,idx):
        img_path = self.img_paths[idx]
        img_path = str(img_path)
        img = cv2.imread(img_path)
        img = img / 255

        mask_path = self.mask_paths[idx]
        mask_path = str(mask_path)
        mask = cv2.imread(mask_path)
        # mask_1lay = mask[:,:,0]
        mask = mask / 255
        # mask = cv2.resize(mask, dsize = None, fx = 0.25, fy = 0.25)
        # 試し用
        # plt.imshow(img)
        # if cpoint.shape != (0,):
        #     plt.scatter(cpoint[:,0],cpoint[:,1], c = "red")
        # if npoint.shape != (0,):    
        #     plt.scatter(npoint[:,0],npoint[:,1], c = "y")
        # plt.show()
        # plt.close()
        # ここまで

        gt = self.gt[idx]
        
        img = torch.from_numpy(img.astype(np.float32))
        mask = torch.from_numpy(mask.astype(np.float32))
        return img.permute(2, 0, 1), img_path, mask.permute(2, 0, 1), gt, idx
        # return img, img_path, mask

class Est_proportion_pred(object):
    def __init__(self,base_path):
        self.img_paths = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_1_PDL122C3").glob("*.png"))
        self.mask_paths = sorted(Path(f"{base_path}/256_data/point_mask/96_1_PDL122C3").glob("*.png"))
    
    def __len__(self):
        return len(self.img_paths)

    def __getitem__(self,idx):
        img_path = self.img_paths[idx]
        img_path = str(img_path)
        img = cv2.imread(img_path)
        img = img / 255

        mask_path = self.mask_paths[idx]
        mask_path = str(mask_path)
        mask = cv2.imread(mask_path)
        # mask_1lay = mask[:,:,0]
        mask = mask / 255
        # mask = cv2.resize(mask, dsize = None, fx = 0.25, fy = 0.25)
        # 試し用
        # plt.imshow(img)
        # if cpoint.shape != (0,):
        #     plt.scatter(cpoint[:,0],cpoint[:,1], c = "red")
        # if npoint.shape != (0,):    
        #     plt.scatter(npoint[:,0],npoint[:,1], c = "y")
        # plt.show()
        # plt.close()
        # ここまで
        
        img = torch.from_numpy(img.astype(np.float32))
        mask = torch.from_numpy(mask.astype(np.float32))
        return img.permute(2, 0, 1), img_path, mask.permute(2, 0, 1), idx
        # return img, img_path, mask


class Est_proportion_train123(object):
    def __init__(self,base_path):
        self.img_paths0 = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_1_PDL122C3").glob("*.png"))
        self.mask_paths0 = sorted(Path(f"{base_path}/256_data/point_mask/96_1_PDL122C3").glob("*.png"))
        self.img_paths1 = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_2_PDL122C3").glob("*.png"))
        self.mask_paths1 = sorted(Path(f"{base_path}/256_data/point_mask/96_2_PDL122C3").glob("*.png"))
        self.img_paths2 = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_3_PDL122C3").glob("*.png"))
        self.mask_paths2 = sorted(Path(f"{base_path}/256_data/point_mask/96_3_PDL122C3").glob("*.png"))

        self.img_paths = self.img_paths0 + self.img_paths1 + self.img_paths2
        self.mask_paths = self.mask_paths0 + self.mask_paths1 + self.mask_paths2

        # over sampling
        # exel sheet 2 numpy array
        input_file_name = '/home/okuo-takumi/pdl1_code/pdl1_rate/eval/delete7/train/96_123_rate.xlsx'
        book = openpyxl.load_workbook(input_file_name)
        sheet = book['Sheet1']
        gt_labels = []
        for idx in range(sheet.max_row - 1):
            gt_val = sheet[f'H{idx+2}'].value
            if gt_val == 1:
                gt_val = 3
            if gt_val == 2:
                gt_val = 3
            gt_labels.append(gt_val)
        gt_labels = np.array(gt_labels)

        # max num of class
        img_path_np = np.array(self.img_paths)
        mask_path_np = np.array(self.mask_paths)

        label_num_list = []
        img_path_per_label = {}
        mask_path_per_label = {}
        for u_id in np.unique(gt_labels):
            label_num_list.append(sum(gt_labels == u_id))
            img_path_per_label[u_id] = img_path_np[gt_labels == u_id]
            mask_path_per_label[u_id] = mask_path_np[gt_labels == u_id]
        num_img = max(label_num_list)
        
        new_img_path = []
        new_mask_path = []
        new_gt = []
        val_img_path = []
        val_mask_path = []
        val_gt = []
        gt = []
        for u_id, num_list in zip(np.unique(gt_labels), label_num_list):
            val_num = int(num_list / 5)
            ns = []
            while len(ns) < val_num:
                n = random.randint(0,len(img_path_per_label[u_id])-1)
                if not n in ns:
                    ns.append(n)
            ns.sort()
            for n in ns:
                val_img_path.append(img_path_per_label[u_id][n])
                val_mask_path.append(mask_path_per_label[u_id][n])
                img_path_per_label[u_id] = np.delete(img_path_per_label[u_id],n)
                mask_path_per_label[u_id] = np.delete(mask_path_per_label[u_id],n)
                for i in range(len(ns)):
                    ns[i] = ns[i] - 1

            val_gt.extend(np.repeat([u_id],val_num))
            gt.extend(np.repeat([u_id],num_list - val_num))
        gt = np.array(gt)


        for u_id, num_list in zip(np.unique(gt), label_num_list):
            x = num_img // num_list
            for idx in range(x):
                new_img_path.extend(img_path_per_label[u_id])
            x = num_img % num_list
            new_img_path.extend(img_path_per_label[u_id][:x])

        for u_id, num_list in zip(np.unique(gt), label_num_list):
            x = num_img // num_list
            for idx in range(x):
                new_mask_path.extend(mask_path_per_label[u_id])
            x = num_img % num_list
            new_mask_path.extend(mask_path_per_label[u_id][:x])

        for u_id, num_list in zip(np.unique(gt), label_num_list):
            x = num_img // num_list
            for idx in range(x):
                new_gt.extend(gt[gt == u_id])
            x = num_img % num_list
            new_gt.extend(gt[gt == u_id][:x])

        self.new_img_path = new_img_path
        self.new_mask_path = new_mask_path
        self.new_gt = new_gt
        self.val_img_path = val_img_path
        self.val_mask_path = val_mask_path
        self.val_gt = val_gt

        

    def __len__(self):
        return len(self.new_img_path)

    def __getitem__(self,idx):
        img_path = self.new_img_path[idx]
        img_path = str(img_path)
        img = cv2.imread(img_path)
        mask_path = self.new_mask_path[idx]
        mask_path = str(mask_path)
        mask = cv2.imread(mask_path)

        rand1 = random.randint(0,3)
        rand2 = random.randint(-2,1)
        if rand1 == 1:
            img = cv2.rotate(img, cv2.ROTATE_90_CLOCKWISE)
            mask = cv2.rotate(mask, cv2.ROTATE_90_CLOCKWISE)
        elif rand1 == 2:
            img = cv2.rotate(img, cv2.ROTATE_180)
            mask = cv2.rotate(mask, cv2.ROTATE_180)
        elif rand1 == 3:
            img = cv2.rotate(img, cv2.ROTATE_90_COUNTERCLOCKWISE)
            mask = cv2.rotate(mask, cv2.ROTATE_90_COUNTERCLOCKWISE)

        if rand2 >= -1:
            img = cv2.flip(img,rand2)
            mask = cv2.flip(mask,rand2)

        img = img / 255
        mask = mask / 255

        gt = self.new_gt[idx]
        
        img = torch.from_numpy(img.astype(np.float32))
        mask = torch.from_numpy(mask.astype(np.float32))
        return img.permute(2, 0, 1), img_path, mask.permute(2, 0, 1), gt, idx

class Est_proportion_val123(object):
    def __init__(self,base_path):
        self.img_paths0 = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_1_PDL122C3").glob("*.png"))
        self.mask_paths0 = sorted(Path(f"{base_path}/256_data/point_mask/96_1_PDL122C3").glob("*.png"))
        self.img_paths1 = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_2_PDL122C3").glob("*.png"))
        self.mask_paths1 = sorted(Path(f"{base_path}/256_data/point_mask/96_2_PDL122C3").glob("*.png"))
        self.img_paths2 = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_3_PDL122C3").glob("*.png"))
        self.mask_paths2 = sorted(Path(f"{base_path}/256_data/point_mask/96_3_PDL122C3").glob("*.png"))

        self.img_paths = self.img_paths0 + self.img_paths1 + self.img_paths2
        self.mask_paths = self.mask_paths0 + self.mask_paths1 + self.mask_paths2

        # over sampling
        # exel sheet 2 numpy array
        input_file_name = '/home/okuo-takumi/pdl1_code/pdl1_rate/eval/delete7/train/96_123_rate.xlsx'
        book = openpyxl.load_workbook(input_file_name)
        sheet = book['Sheet1']
        gt_labels = []
        for idx in range(sheet.max_row - 1):
            gt_val = sheet[f'H{idx+2}'].value
            if gt_val == 1:
                gt_val = 3
            if gt_val == 2:
                gt_val = 3
            gt_labels.append(gt_val)
        gt_labels = np.array(gt_labels)
        self.gt = gt_labels
        

    def __len__(self):
        return len(self.img_paths)

    def __getitem__(self,idx):
        img_path = self.img_paths[idx]
        img_path = str(img_path)
        img = cv2.imread(img_path)
        img = img / 255

        mask_path = self.mask_paths[idx]
        mask_path = str(mask_path)
        mask = cv2.imread(mask_path)
        # mask_1lay = mask[:,:,0]
        mask = mask / 255
        # mask = cv2.resize(mask, dsize = None, fx = 0.25, fy = 0.25)
        # 試し用
        # plt.imshow(img)
        # if cpoint.shape != (0,):
        #     plt.scatter(cpoint[:,0],cpoint[:,1], c = "red")
        # if npoint.shape != (0,):    
        #     plt.scatter(npoint[:,0],npoint[:,1], c = "y")
        # plt.show()
        # plt.close()
        # ここまで

        gt = self.gt[idx]
        
        img = torch.from_numpy(img.astype(np.float32))
        mask = torch.from_numpy(mask.astype(np.float32))
        return img.permute(2, 0, 1), img_path, mask.permute(2, 0, 1), gt, idx
        # return img, img_path, mask

    
class Est_proportion_train124(object):
    def __init__(self,base_path):
        self.img_paths0 = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_1_PDL122C3").glob("*.png"))
        self.mask_paths0 = sorted(Path(f"{base_path}/256_data/point_mask/96_1_PDL122C3").glob("*.png"))
        self.img_paths1 = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_2_PDL122C3").glob("*.png"))
        self.mask_paths1 = sorted(Path(f"{base_path}/256_data/point_mask/96_2_PDL122C3").glob("*.png"))
        self.img_paths2 = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_4_PDL122C3").glob("*.png"))
        self.mask_paths2 = sorted(Path(f"{base_path}/256_data/point_mask/96_4_PDL122C3").glob("*.png"))

        self.img_paths = self.img_paths0 + self.img_paths1 + self.img_paths2
        self.mask_paths = self.mask_paths0 + self.mask_paths1 + self.mask_paths2

        # over sampling
        # exel sheet 2 numpy array
        input_file_name = '/home/okuo-takumi/pdl1_code/pdl1_rate/eval/delete7/train/96_124_rate.xlsx'
        book = openpyxl.load_workbook(input_file_name)
        sheet = book['Sheet1']
        gt_labels = []
        for idx in range(sheet.max_row - 1):
            gt_val = sheet[f'H{idx+2}'].value
            if gt_val == 1:
                gt_val = 3
            if gt_val == 2:
                gt_val = 3
            gt_labels.append(gt_val)
        gt_labels = np.array(gt_labels)

        # max num of class
        img_path_np = np.array(self.img_paths)
        mask_path_np = np.array(self.mask_paths)

        label_num_list = []
        img_path_per_label = {}
        mask_path_per_label = {}
        for u_id in np.unique(gt_labels):
            label_num_list.append(sum(gt_labels == u_id))
            img_path_per_label[u_id] = img_path_np[gt_labels == u_id]
            mask_path_per_label[u_id] = mask_path_np[gt_labels == u_id]
        num_img = max(label_num_list)
        
        new_img_path = []
        new_mask_path = []
        new_gt = []

        for u_id, num_list in zip(np.unique(gt_labels), label_num_list):
            x = num_img // num_list
            for idx in range(x):
                new_img_path.extend(img_path_per_label[u_id])
            x = num_img % num_list
            new_img_path.extend(img_path_per_label[u_id][:x])

        for u_id, num_list in zip(np.unique(gt_labels), label_num_list):
            x = num_img // num_list
            for idx in range(x):
                new_mask_path.extend(mask_path_per_label[u_id])
            x = num_img % num_list
            new_mask_path.extend(mask_path_per_label[u_id][:x])

        for u_id, num_list in zip(np.unique(gt_labels), label_num_list):
            x = num_img // num_list
            for idx in range(x):
                new_gt.extend(gt_labels[gt_labels == u_id])
            x = num_img % num_list
            new_gt.extend(gt_labels[gt_labels == u_id][:x])

        self.new_img_path = new_img_path
        self.new_mask_path = new_mask_path
        self.new_gt = new_gt

        

    def __len__(self):
        return len(self.new_img_path)

    def __getitem__(self,idx):
        img_path = self.new_img_path[idx]
        img_path = str(img_path)
        img = cv2.imread(img_path)
        mask_path = self.new_mask_path[idx]
        mask_path = str(mask_path)
        mask = cv2.imread(mask_path)

        rand1 = random.randint(0,3)
        rand2 = random.randint(-2,1)
        if rand1 == 1:
            img = cv2.rotate(img, cv2.ROTATE_90_CLOCKWISE)
            mask = cv2.rotate(mask, cv2.ROTATE_90_CLOCKWISE)
        elif rand1 == 2:
            img = cv2.rotate(img, cv2.ROTATE_180)
            mask = cv2.rotate(mask, cv2.ROTATE_180)
        elif rand1 == 3:
            img = cv2.rotate(img, cv2.ROTATE_90_COUNTERCLOCKWISE)
            mask = cv2.rotate(mask, cv2.ROTATE_90_COUNTERCLOCKWISE)

        if rand2 >= -1:
            img = cv2.flip(img,rand2)
            mask = cv2.flip(mask,rand2)

        img = img / 255
        mask = mask / 255

        gt = self.new_gt[idx]
        
        img = torch.from_numpy(img.astype(np.float32))
        mask = torch.from_numpy(mask.astype(np.float32))
        return img.permute(2, 0, 1), img_path, mask.permute(2, 0, 1), gt, idx
        # return img, img_path, mask

class Est_proportion_val124(object):
    def __init__(self,base_path):
        self.img_paths0 = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_1_PDL122C3").glob("*.png"))
        self.mask_paths0 = sorted(Path(f"{base_path}/256_data/point_mask/96_1_PDL122C3").glob("*.png"))
        self.img_paths1 = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_2_PDL122C3").glob("*.png"))
        self.mask_paths1 = sorted(Path(f"{base_path}/256_data/point_mask/96_2_PDL122C3").glob("*.png"))
        self.img_paths2 = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_4_PDL122C3").glob("*.png"))
        self.mask_paths2 = sorted(Path(f"{base_path}/256_data/point_mask/96_4_PDL122C3").glob("*.png"))

        self.img_paths = self.img_paths0 + self.img_paths1 + self.img_paths2
        self.mask_paths = self.mask_paths0 + self.mask_paths1 + self.mask_paths2

        # over sampling
        # exel sheet 2 numpy array
        input_file_name = '/home/okuo-takumi/pdl1_code/pdl1_rate/eval/delete7/train/96_124_rate.xlsx'
        book = openpyxl.load_workbook(input_file_name)
        sheet = book['Sheet1']
        gt_labels = []
        for idx in range(sheet.max_row - 1):
            gt_val = sheet[f'H{idx+2}'].value
            if gt_val == 1:
                gt_val = 3
            if gt_val == 2:
                gt_val = 3
            gt_labels.append(gt_val)
        gt_labels = np.array(gt_labels)
        self.gt = gt_labels
        

    def __len__(self):
        return len(self.img_paths)

    def __getitem__(self,idx):
        img_path = self.img_paths[idx]
        img_path = str(img_path)
        img = cv2.imread(img_path)
        img = img / 255

        mask_path = self.mask_paths[idx]
        mask_path = str(mask_path)
        mask = cv2.imread(mask_path)
        # mask_1lay = mask[:,:,0]
        mask = mask / 255
        # mask = cv2.resize(mask, dsize = None, fx = 0.25, fy = 0.25)
        # 試し用
        # plt.imshow(img)
        # if cpoint.shape != (0,):
        #     plt.scatter(cpoint[:,0],cpoint[:,1], c = "red")
        # if npoint.shape != (0,):    
        #     plt.scatter(npoint[:,0],npoint[:,1], c = "y")
        # plt.show()
        # plt.close()
        # ここまで

        gt = self.gt[idx]
        
        img = torch.from_numpy(img.astype(np.float32))
        mask = torch.from_numpy(mask.astype(np.float32))
        return img.permute(2, 0, 1), img_path, mask.permute(2, 0, 1), gt, idx
        # return img, img_path, mask

class Est_proportion_train134(object):
    def __init__(self,base_path):
        self.img_paths0 = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_1_PDL122C3").glob("*.png"))
        self.mask_paths0 = sorted(Path(f"{base_path}/256_data/point_mask/96_1_PDL122C3").glob("*.png"))
        self.img_paths1 = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_3_PDL122C3").glob("*.png"))
        self.mask_paths1 = sorted(Path(f"{base_path}/256_data/point_mask/96_3_PDL122C3").glob("*.png"))
        self.img_paths2 = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_4_PDL122C3").glob("*.png"))
        self.mask_paths2 = sorted(Path(f"{base_path}/256_data/point_mask/96_4_PDL122C3").glob("*.png"))

        self.img_paths = self.img_paths0 + self.img_paths1 + self.img_paths2
        self.mask_paths = self.mask_paths0 + self.mask_paths1 + self.mask_paths2

        # over sampling
        # exel sheet 2 numpy array
        input_file_name = '/home/okuo-takumi/pdl1_code/pdl1_rate/eval/delete7/train/96_134_rate.xlsx'
        book = openpyxl.load_workbook(input_file_name)
        sheet = book['Sheet1']
        gt_labels = []
        for idx in range(sheet.max_row - 1):
            gt_val = sheet[f'H{idx+2}'].value
            if gt_val == 1:
                gt_val = 3
            if gt_val == 2:
                gt_val = 3
            gt_labels.append(gt_val)
        gt_labels = np.array(gt_labels)

        # max num of class
        img_path_np = np.array(self.img_paths)
        mask_path_np = np.array(self.mask_paths)

        label_num_list = []
        img_path_per_label = {}
        mask_path_per_label = {}
        for u_id in np.unique(gt_labels):
            label_num_list.append(sum(gt_labels == u_id))
            img_path_per_label[u_id] = img_path_np[gt_labels == u_id]
            mask_path_per_label[u_id] = mask_path_np[gt_labels == u_id]
        num_img = max(label_num_list)
        
        new_img_path = []
        new_mask_path = []
        new_gt = []

        for u_id, num_list in zip(np.unique(gt_labels), label_num_list):
            x = num_img // num_list
            for idx in range(x):
                new_img_path.extend(img_path_per_label[u_id])
            x = num_img % num_list
            new_img_path.extend(img_path_per_label[u_id][:x])

        for u_id, num_list in zip(np.unique(gt_labels), label_num_list):
            x = num_img // num_list
            for idx in range(x):
                new_mask_path.extend(mask_path_per_label[u_id])
            x = num_img % num_list
            new_mask_path.extend(mask_path_per_label[u_id][:x])

        for u_id, num_list in zip(np.unique(gt_labels), label_num_list):
            x = num_img // num_list
            for idx in range(x):
                new_gt.extend(gt_labels[gt_labels == u_id])
            x = num_img % num_list
            new_gt.extend(gt_labels[gt_labels == u_id][:x])

        self.new_img_path = new_img_path
        self.new_mask_path = new_mask_path
        self.new_gt = new_gt

        

    def __len__(self):
        return len(self.new_img_path)

    def __getitem__(self,idx):
        img_path = self.new_img_path[idx]
        img_path = str(img_path)
        img = cv2.imread(img_path)
        mask_path = self.new_mask_path[idx]
        mask_path = str(mask_path)
        mask = cv2.imread(mask_path)

        rand1 = random.randint(0,3)
        rand2 = random.randint(-2,1)
        if rand1 == 1:
            img = cv2.rotate(img, cv2.ROTATE_90_CLOCKWISE)
            mask = cv2.rotate(mask, cv2.ROTATE_90_CLOCKWISE)
        elif rand1 == 2:
            img = cv2.rotate(img, cv2.ROTATE_180)
            mask = cv2.rotate(mask, cv2.ROTATE_180)
        elif rand1 == 3:
            img = cv2.rotate(img, cv2.ROTATE_90_COUNTERCLOCKWISE)
            mask = cv2.rotate(mask, cv2.ROTATE_90_COUNTERCLOCKWISE)

        if rand2 >= -1:
            img = cv2.flip(img,rand2)
            mask = cv2.flip(mask,rand2)

        img = img / 255
        mask = mask / 255

        gt = self.new_gt[idx]
        
        img = torch.from_numpy(img.astype(np.float32))
        mask = torch.from_numpy(mask.astype(np.float32))
        return img.permute(2, 0, 1), img_path, mask.permute(2, 0, 1), gt, idx
        # return img, img_path, mask

class Est_proportion_val134(object):
    def __init__(self,base_path):
        self.img_paths0 = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_1_PDL122C3").glob("*.png"))
        self.mask_paths0 = sorted(Path(f"{base_path}/256_data/point_mask/96_1_PDL122C3").glob("*.png"))
        self.img_paths1 = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_3_PDL122C3").glob("*.png"))
        self.mask_paths1 = sorted(Path(f"{base_path}/256_data/point_mask/96_3_PDL122C3").glob("*.png"))
        self.img_paths2 = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_4_PDL122C3").glob("*.png"))
        self.mask_paths2 = sorted(Path(f"{base_path}/256_data/point_mask/96_4_PDL122C3").glob("*.png"))

        self.img_paths = self.img_paths0 + self.img_paths1 + self.img_paths2
        self.mask_paths = self.mask_paths0 + self.mask_paths1 + self.mask_paths2

        # over sampling
        # exel sheet 2 numpy array
        input_file_name = '/home/okuo-takumi/pdl1_code/pdl1_rate/eval/delete7/train/96_134_rate.xlsx'
        book = openpyxl.load_workbook(input_file_name)
        sheet = book['Sheet1']
        gt_labels = []
        for idx in range(sheet.max_row - 1):
            gt_val = sheet[f'H{idx+2}'].value
            if gt_val == 1:
                gt_val = 3
            if gt_val == 2:
                gt_val = 3
            gt_labels.append(gt_val)
        gt_labels = np.array(gt_labels)
        self.gt = gt_labels
        

    def __len__(self):
        return len(self.img_paths)

    def __getitem__(self,idx):
        img_path = self.img_paths[idx]
        img_path = str(img_path)
        img = cv2.imread(img_path)
        img = img / 255

        mask_path = self.mask_paths[idx]
        mask_path = str(mask_path)
        mask = cv2.imread(mask_path)
        # mask_1lay = mask[:,:,0]
        mask = mask / 255
        # mask = cv2.resize(mask, dsize = None, fx = 0.25, fy = 0.25)
        # 試し用
        # plt.imshow(img)
        # if cpoint.shape != (0,):
        #     plt.scatter(cpoint[:,0],cpoint[:,1], c = "red")
        # if npoint.shape != (0,):    
        #     plt.scatter(npoint[:,0],npoint[:,1], c = "y")
        # plt.show()
        # plt.close()
        # ここまで

        gt = self.gt[idx]
        
        img = torch.from_numpy(img.astype(np.float32))
        mask = torch.from_numpy(mask.astype(np.float32))
        return img.permute(2, 0, 1), img_path, mask.permute(2, 0, 1), gt, idx
        # return img, img_path, mask


class Est_proportion_train234(object):
    def __init__(self,base_path):
        self.img_paths0 = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_2_PDL122C3").glob("*.png"))
        self.mask_paths0 = sorted(Path(f"{base_path}/256_data/point_mask/96_2_PDL122C3").glob("*.png"))
        self.img_paths1 = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_3_PDL122C3").glob("*.png"))
        self.mask_paths1 = sorted(Path(f"{base_path}/256_data/point_mask/96_3_PDL122C3").glob("*.png"))
        self.img_paths2 = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_4_PDL122C3").glob("*.png"))
        self.mask_paths2 = sorted(Path(f"{base_path}/256_data/point_mask/96_4_PDL122C3").glob("*.png"))

        self.img_paths = self.img_paths0 + self.img_paths1 + self.img_paths2
        self.mask_paths = self.mask_paths0 + self.mask_paths1 + self.mask_paths2

        # over sampling
        # exel sheet 2 numpy array
        input_file_name = '/home/okuo-takumi/pdl1_code/pdl1_rate/eval/delete7/train/96_234_rate.xlsx'
        book = openpyxl.load_workbook(input_file_name)
        sheet = book['Sheet1']
        gt_labels = []
        for idx in range(sheet.max_row - 1):
            gt_val = sheet[f'H{idx+2}'].value
            if gt_val == 1:
                gt_val = 3
            if gt_val == 2:
                gt_val = 3
            gt_labels.append(gt_val)
        gt_labels = np.array(gt_labels)

        # max num of class
        img_path_np = np.array(self.img_paths)
        mask_path_np = np.array(self.mask_paths)

        label_num_list = []
        img_path_per_label = {}
        mask_path_per_label = {}
        for u_id in np.unique(gt_labels):
            label_num_list.append(sum(gt_labels == u_id))
            img_path_per_label[u_id] = img_path_np[gt_labels == u_id]
            mask_path_per_label[u_id] = mask_path_np[gt_labels == u_id]
        num_img = max(label_num_list)
        
        new_img_path = []
        new_mask_path = []
        new_gt = []

        for u_id, num_list in zip(np.unique(gt_labels), label_num_list):
            x = num_img // num_list
            for idx in range(x):
                new_img_path.extend(img_path_per_label[u_id])
            x = num_img % num_list
            new_img_path.extend(img_path_per_label[u_id][:x])

        for u_id, num_list in zip(np.unique(gt_labels), label_num_list):
            x = num_img // num_list
            for idx in range(x):
                new_mask_path.extend(mask_path_per_label[u_id])
            x = num_img % num_list
            new_mask_path.extend(mask_path_per_label[u_id][:x])

        for u_id, num_list in zip(np.unique(gt_labels), label_num_list):
            x = num_img // num_list
            for idx in range(x):
                new_gt.extend(gt_labels[gt_labels == u_id])
            x = num_img % num_list
            new_gt.extend(gt_labels[gt_labels == u_id][:x])

        self.new_img_path = new_img_path
        self.new_mask_path = new_mask_path
        self.new_gt = new_gt

        

    def __len__(self):
        return len(self.new_img_path)

    def __getitem__(self,idx):
        img_path = self.new_img_path[idx]
        img_path = str(img_path)
        img = cv2.imread(img_path)
        mask_path = self.new_mask_path[idx]
        mask_path = str(mask_path)
        mask = cv2.imread(mask_path)

        rand1 = random.randint(0,3)
        rand2 = random.randint(-2,1)
        if rand1 == 1:
            img = cv2.rotate(img, cv2.ROTATE_90_CLOCKWISE)
            mask = cv2.rotate(mask, cv2.ROTATE_90_CLOCKWISE)
        elif rand1 == 2:
            img = cv2.rotate(img, cv2.ROTATE_180)
            mask = cv2.rotate(mask, cv2.ROTATE_180)
        elif rand1 == 3:
            img = cv2.rotate(img, cv2.ROTATE_90_COUNTERCLOCKWISE)
            mask = cv2.rotate(mask, cv2.ROTATE_90_COUNTERCLOCKWISE)

        if rand2 >= -1:
            img = cv2.flip(img,rand2)
            mask = cv2.flip(mask,rand2)

        img = img / 255
        mask = mask / 255

        gt = self.new_gt[idx]
        
        img = torch.from_numpy(img.astype(np.float32))
        mask = torch.from_numpy(mask.astype(np.float32))
        return img.permute(2, 0, 1), img_path, mask.permute(2, 0, 1), gt, idx
        # return img, img_path, mask

class Est_proportion_val234(object):
    def __init__(self,base_path):
        self.img_paths0 = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_2_PDL122C3").glob("*.png"))
        self.mask_paths0 = sorted(Path(f"{base_path}/256_data/point_mask/96_2_PDL122C3").glob("*.png"))
        self.img_paths1 = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_3_PDL122C3").glob("*.png"))
        self.mask_paths1 = sorted(Path(f"{base_path}/256_data/point_mask/96_3_PDL122C3").glob("*.png"))
        self.img_paths2 = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_4_PDL122C3").glob("*.png"))
        self.mask_paths2 = sorted(Path(f"{base_path}/256_data/point_mask/96_4_PDL122C3").glob("*.png"))

        self.img_paths = self.img_paths0 + self.img_paths1 + self.img_paths2
        self.mask_paths = self.mask_paths0 + self.mask_paths1 + self.mask_paths2

        # over sampling
        # exel sheet 2 numpy array
        input_file_name = '/home/okuo-takumi/pdl1_code/pdl1_rate/eval/delete7/train/96_234_rate.xlsx'
        book = openpyxl.load_workbook(input_file_name)
        sheet = book['Sheet1']
        gt_labels = []
        for idx in range(sheet.max_row - 1):
            gt_val = sheet[f'H{idx+2}'].value
            if gt_val == 1:
                gt_val = 3
            if gt_val == 2:
                gt_val = 3
            gt_labels.append(gt_val)
        gt_labels = np.array(gt_labels)
        self.gt = gt_labels
        

    def __len__(self):
        return len(self.img_paths)

    def __getitem__(self,idx):
        img_path = self.img_paths[idx]
        img_path = str(img_path)
        img = cv2.imread(img_path)
        img = img / 255

        mask_path = self.mask_paths[idx]
        mask_path = str(mask_path)
        mask = cv2.imread(mask_path)
        # mask_1lay = mask[:,:,0]
        mask = mask / 255
        # mask = cv2.resize(mask, dsize = None, fx = 0.25, fy = 0.25)
        # 試し用
        # plt.imshow(img)
        # if cpoint.shape != (0,):
        #     plt.scatter(cpoint[:,0],cpoint[:,1], c = "red")
        # if npoint.shape != (0,):    
        #     plt.scatter(npoint[:,0],npoint[:,1], c = "y")
        # plt.show()
        # plt.close()
        # ここまで

        gt = self.gt[idx]
        
        img = torch.from_numpy(img.astype(np.float32))
        mask = torch.from_numpy(mask.astype(np.float32))
        return img.permute(2, 0, 1), img_path, mask.permute(2, 0, 1), gt, idx
        # return img, img_path, mask


class Est_proportion_pred1(object):
    def __init__(self,base_path):
        self.img_paths = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_1_PDL122C3").glob("*.png"))
        self.mask_paths = sorted(Path(f"{base_path}/256_data/point_mask/96_1_PDL122C3").glob("*.png"))
    
    def __len__(self):
        return len(self.img_paths)

    def __getitem__(self,idx):
        img_path = self.img_paths[idx]
        img_path = str(img_path)
        img = cv2.imread(img_path)
        img = img / 255

        mask_path = self.mask_paths[idx]
        mask_path = str(mask_path)
        mask = cv2.imread(mask_path)
        mask = mask / 255
        
        img = torch.from_numpy(img.astype(np.float32))
        mask = torch.from_numpy(mask.astype(np.float32))
        return img.permute(2, 0, 1), img_path, mask.permute(2, 0, 1), idx

class Est_proportion_pred2(object):
    def __init__(self,base_path):
        self.img_paths = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_2_PDL122C3").glob("*.png"))
        self.mask_paths = sorted(Path(f"{base_path}/256_data/point_mask/96_2_PDL122C3").glob("*.png"))
    
    def __len__(self):
        return len(self.img_paths)

    def __getitem__(self,idx):
        img_path = self.img_paths[idx]
        img_path = str(img_path)
        img = cv2.imread(img_path)
        img = img / 255

        mask_path = self.mask_paths[idx]
        mask_path = str(mask_path)
        mask = cv2.imread(mask_path)
        mask = mask / 255
        
        img = torch.from_numpy(img.astype(np.float32))
        mask = torch.from_numpy(mask.astype(np.float32))
        return img.permute(2, 0, 1), img_path, mask.permute(2, 0, 1), idx

class Est_proportion_pred3(object):
    def __init__(self,base_path):
        self.img_paths = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_3_PDL122C3").glob("*.png"))
        self.mask_paths = sorted(Path(f"{base_path}/256_data/point_mask/96_3_PDL122C3").glob("*.png"))
    
    def __len__(self):
        return len(self.img_paths)

    def __getitem__(self,idx):
        img_path = self.img_paths[idx]
        img_path = str(img_path)
        img = cv2.imread(img_path)
        img = img / 255

        mask_path = self.mask_paths[idx]
        mask_path = str(mask_path)
        mask = cv2.imread(mask_path)
        mask = mask / 255
        
        img = torch.from_numpy(img.astype(np.float32))
        mask = torch.from_numpy(mask.astype(np.float32))
        return img.permute(2, 0, 1), img_path, mask.permute(2, 0, 1), idx

class Est_proportion_pred4(object):
    def __init__(self,base_path):
        self.img_paths = sorted(Path(f"{base_path}/1024_data/img_mask/r=10/96_4_PDL122C3").glob("*.png"))
        self.mask_paths = sorted(Path(f"{base_path}/256_data/point_mask/96_4_PDL122C3").glob("*.png"))
    
    def __len__(self):
        return len(self.img_paths)

    def __getitem__(self,idx):
        img_path = self.img_paths[idx]
        img_path = str(img_path)
        img = cv2.imread(img_path)
        img = img / 255

        mask_path = self.mask_paths[idx]
        mask_path = str(mask_path)
        mask = cv2.imread(mask_path)
        mask = mask / 255
        
        img = torch.from_numpy(img.astype(np.float32))
        mask = torch.from_numpy(mask.astype(np.float32))
        return img.permute(2, 0, 1), img_path, mask.permute(2, 0, 1), idx

class Train(object):
    def __init__(self,base_path):
        self.img_paths = sorted(Path(f"{base_path}/resize_data/2048").glob("*.png"))
        self.mask_paths = sorted(Path(f"{base_path}/mask/512").glob("*.png"))
    
    def __len__(self):
        return int(len(self.img_paths))
    
    def __getitem__(self,idx):
        img_path = self.img_paths[idx]
        img_path = str(img_path)
        img = cv2.imread(img_path)

        mask_path = self.mask_paths[idx]
        mask_path = str(mask_path)
        mask = cv2.imread(mask_path)
        
        #augmentation
        rand1 = random.randint(0,3)
        rand2 = random.randint(-2,1)
        if rand1 == 1:
            img = cv2.rotate(img, cv2.ROTATE_90_CLOCKWISE)
            mask = cv2.rotate(mask, cv2.ROTATE_90_CLOCKWISE)
        elif rand1 == 2:
            img = cv2.rotate(img, cv2.ROTATE_180)
            mask = cv2.rotate(mask, cv2.ROTATE_180)
        elif rand1 == 3:
            img = cv2.rotate(img, cv2.ROTATE_90_COUNTERCLOCKWISE)
            mask = cv2.rotate(mask, cv2.ROTATE_90_COUNTERCLOCKWISE)

        if rand2 >= -1:
            img = cv2.flip(img,rand2)
            mask = cv2.flip(mask,rand2)

        img = img / 255
        mask = mask / 255

        gt = int(img_path[-5])
        
        img = torch.from_numpy(img.astype(np.float32))
        mask = torch.from_numpy(mask.astype(np.float32))
        return img.permute(2, 0, 1), img_path, mask.permute(2, 0, 1), gt, idx



class Pred(object):
    def __init__(self,base_path):
        self.img_paths = sorted(Path(f"{base_path}/resize_data/2048").glob("*.png"))
        self.mask_paths = sorted(Path(f"{base_path}/mask/512").glob("*.png"))
    
    def __len__(self):
        return len(self.img_paths)

    def __getitem__(self,idx):
        img_path = self.img_paths[idx]
        img_path = str(img_path)
        img = cv2.imread(img_path)
        img = img / 255

        mask_path = self.mask_paths[idx]
        mask_path = str(mask_path)
        mask = cv2.imread(mask_path)
        mask = mask / 255

        gt = int(img_path[-5])

        img = torch.from_numpy(img.astype(np.float32))
        mask = torch.from_numpy(mask.astype(np.float32))
        return img.permute(2, 0, 1), img_path, mask.permute(2, 0, 1), gt, idx